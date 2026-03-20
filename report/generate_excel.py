#!/usr/bin/env python3
"""AI自動法廷システム 実験結果Excel生成スクリプト

outputs_agent/ 以下の全 run_* ディレクトリを読み込み、
results_summary.xlsx を生成する。

構成:
  - Sheet 1「実行一覧」: 全事案の横断比較（1行1事案）
  - Sheet 2〜N「事案N_<短縮名>」: 事案ごとに弁論経過・争点整理・合議・判決・批評を1シートに集約

再実行すれば新しいrunも自動的に追加される。
"""

import json
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# ── 設定 ──
BASE_DIR = Path(__file__).resolve().parent.parent
OUTPUTS_DIR = BASE_DIR / "outputs_agent"
OUTPUT_FILE = Path(__file__).resolve().parent / "results_summary.xlsx"

# ── スタイル ──
HEADER_FONT = Font(name="游ゴシック", bold=True, size=10, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)

SECTION_FONT = Font(name="游ゴシック", bold=True, size=11, color="FFFFFF")
SECTION_FILL = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")

SUBHEADER_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
SUBHEADER_FONT = Font(name="游ゴシック", bold=True, size=10)

CELL_FONT = Font(name="游ゴシック", size=9)
CELL_ALIGN = Alignment(vertical="top", wrap_text=True)
LABEL_FONT = Font(name="游ゴシック", bold=True, size=9)

THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

JUDGE_NAMES = {
    "associate_judge_1": "陪席裁判官1（法的安定性）",
    "associate_judge_2": "陪席裁判官2（具体的妥当性）",
    "presiding_judge": "裁判長",
}


# ── ユーティリティ ──

def load_json(path: Path):
    if path.exists():
        with open(path, "r", encoding="utf-8") as f:
            return json.load(f)
    return None


def write_cell(ws, row, col, value, font=None, alignment=None, fill=None):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = font or CELL_FONT
    cell.alignment = alignment or CELL_ALIGN
    cell.border = THIN_BORDER
    if fill:
        cell.fill = fill
    return cell


def write_row(ws, row, values, col_widths=None, font=None, alignment=None, fill=None):
    for col_idx, val in enumerate(values, 1):
        write_cell(ws, row, col_idx, val, font=font, alignment=alignment, fill=fill)
    if col_widths:
        _auto_row_height(ws, row, col_widths, values)


def set_col_widths(ws, col_widths):
    for col_idx, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def write_header(ws, row, headers, col_widths, num_cols=None):
    """テーブルヘッダ行を書き込む。"""
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER


def write_section_title(ws, row, title, num_cols):
    """セクション見出し行（暗い青帯）を書き込む。"""
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=num_cols)
    cell = ws.cell(row=row, column=1, value=title)
    cell.font = SECTION_FONT
    cell.fill = SECTION_FILL
    cell.alignment = Alignment(horizontal="left", vertical="center")
    cell.border = THIN_BORDER
    for c in range(2, num_cols + 1):
        sc = ws.cell(row=row, column=c)
        sc.fill = SECTION_FILL
        sc.border = THIN_BORDER
    ws.row_dimensions[row].height = 22


def write_subheader(ws, row, label, num_cols):
    """サブヘッダ行（薄い青帯）を書き込む。"""
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=num_cols)
    cell = ws.cell(row=row, column=1, value=label)
    cell.font = SUBHEADER_FONT
    cell.fill = SUBHEADER_FILL
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    cell.border = THIN_BORDER
    for c in range(2, num_cols + 1):
        sc = ws.cell(row=row, column=c)
        sc.fill = SUBHEADER_FILL
        sc.border = THIN_BORDER


def _auto_row_height(ws, row, col_widths, values):
    """長文セルに合わせて行の高さを自動調整する（概算）。"""
    max_lines = 1
    for idx, val in enumerate(values):
        if not isinstance(val, str) or not val:
            continue
        col_w = col_widths[idx] if idx < len(col_widths) else 20
        chars_per_line = max(int(col_w * 1.8), 10)
        lines = 0
        for paragraph in val.split("\n"):
            lines += max(1, -(-len(paragraph) // chars_per_line))
        max_lines = max(max_lines, lines)
    ws.row_dimensions[row].height = min(max(max_lines * 13, 15), 400)


def parse_run_id(run_dir_name: str):
    try:
        ts = run_dir_name.replace("run_", "")
        return datetime.strptime(ts, "%Y%m%d_%H%M%S")
    except ValueError:
        return None


def collect_runs():
    runs = []
    if not OUTPUTS_DIR.exists():
        return runs
    for d in sorted(OUTPUTS_DIR.iterdir()):
        if d.is_dir() and d.name.startswith("run_"):
            runs.append(d)
    return runs


def check_critique_addressed(critique_problem: str, final_reasoning: str) -> str:
    keywords_map = {
        "借地借家法": "借地借家法",
        "賃貸人地位承継": "賃貸人地位",
        "善意・悪意": "善意",
        "176条・177条": "176条",
        "前提事実": "前提事実",
        "信義則": "信義則",
        "解除権": "解除",
        "請求範囲": "部分",
        "背信的悪意者": "背信的悪意者",
        "三段論法": "三段論法",
        "605条": "605条",
        "社会通念": "社会通念",
        "権利濫用": "権利濫用",
    }
    for key, search_term in keywords_map.items():
        if key in critique_problem and search_term in final_reasoning:
            return "対応済"
    return "要確認"


# ================================================================
# データ抽出
# ================================================================
def extract_run_data(run_dir: Path) -> dict:
    run_id = run_dir.name
    run_dt = parse_run_id(run_id)

    meta = load_json(run_dir / "meta.json") or {}
    case_record = load_json(run_dir / "case_record.json") or {}
    issues_current = load_json(run_dir / "issue_table_current.json") or []
    turn_log = load_json(run_dir / "turn_log.json") or []
    round_summaries = load_json(run_dir / "round_summaries.json") or []
    judge_decisions = load_json(run_dir / "judge_decisions.json") or []
    deliberations = load_json(run_dir / "deliberations.json") or []
    critique_log = load_json(run_dir / "critique_log.json") or {}
    draft_judgment = load_json(run_dir / "draft_judgment.json") or {}
    final_judgment = load_json(run_dir / "final_judgment.json") or {}

    parties = case_record.get("parties", [])
    plaintiff = next((p["name"] for p in parties if p["role"] == "plaintiff"), "")
    defendant = next((p["name"] for p in parties if p["role"] == "defendant"), "")

    conclusion = final_judgment.get("conclusion", "")

    all_statutes = []
    for iss in final_judgment.get("issues", []):
        all_statutes.extend(iss.get("statutes", []))
    main_statutes = ", ".join(sorted(set(all_statutes)))

    delib_opinions = []
    for delib in deliberations:
        for op in delib.get("opinions", []):
            delib_opinions.append(op)

    criticisms_raw = critique_log.get("criticisms", [])
    full_final_reasoning = " ".join(
        i.get("reasoning", "") for i in final_judgment.get("issues", [])
    )
    crit_data = []
    for c in criticisms_raw:
        addressed = check_critique_addressed(c.get("problem", ""), full_final_reasoning)
        crit_data.append({
            "target_issue_id": c.get("target_issue_id", ""),
            "authored_by": c.get("authored_by", ""),
            "problem": c.get("problem", ""),
            "reason": c.get("reason", ""),
            "suggestion": c.get("suggestion", ""),
            "addressed": addressed,
        })

    return {
        "run_id": run_id,
        "run_datetime": run_dt.strftime("%Y-%m-%d %H:%M") if run_dt else "",
        "model": "gpt-5",
        "case_title": case_record.get("title", ""),
        "case_summary": case_record.get("case_summary", ""),
        "plaintiff": plaintiff,
        "defendant": defendant,
        "max_rounds": meta.get("max_rounds", ""),
        "actual_rounds": meta.get("current_round", ""),
        "issue_count": len(issues_current),
        "conclusion": conclusion,
        "main_statutes": main_statutes,
        "critique_count": len(criticisms_raw),
        "issues": issues_current,
        "turns": turn_log,
        "round_summaries": round_summaries,
        "judge_decisions": judge_decisions,
        "deliberation_opinions": delib_opinions,
        "draft_judgment": draft_judgment,
        "final_judgment": final_judgment,
        "criticisms": crit_data,
    }


# ================================================================
# Sheet 1: 実行一覧（横断比較）
# ================================================================
def build_overview_sheet(wb, all_runs_data):
    ws = wb.active
    ws.title = "実行一覧"
    headers = [
        "No.", "run_id", "実行日時", "モデル", "事件名",
        "原告", "被告", "max_rounds", "実際ラウンド数",
        "争点数", "判決主文", "主な適用条文", "批評項目数",
        "事案概要",
    ]
    col_widths = [5, 25, 18, 10, 30, 8, 8, 11, 11, 8, 40, 30, 10, 60]
    write_header(ws, 1, headers, col_widths)
    set_col_widths(ws, col_widths)
    ws.freeze_panes = "A2"

    for row_idx, rd in enumerate(all_runs_data, 2):
        values = [
            row_idx - 1,
            rd["run_id"],
            rd["run_datetime"],
            rd.get("model", "gpt-5"),
            rd.get("case_title", ""),
            rd.get("plaintiff", ""),
            rd.get("defendant", ""),
            rd.get("max_rounds", ""),
            rd.get("actual_rounds", ""),
            rd.get("issue_count", 0),
            rd.get("conclusion", ""),
            rd.get("main_statutes", ""),
            rd.get("critique_count", 0),
            rd.get("case_summary", ""),
        ]
        write_row(ws, row_idx, values, col_widths)


# ================================================================
# 事案別シート: 弁論経過・争点整理・合議・判決・批評を1シートに集約
# ================================================================
NUM_COLS = 8  # シート内の最大列数


def build_case_sheet(wb, case_no, rd):
    """1事案分のシートを作成する。"""
    # シート名（Excelの31文字制限に対応）
    title_short = rd.get("case_title", rd["run_id"])[:15]
    sheet_name = f"事案{case_no}_{title_short}"
    if len(sheet_name) > 31:
        sheet_name = sheet_name[:31]
    ws = wb.create_sheet(sheet_name)

    col_widths = [15, 15, 15, 25, 20, 60, 60, 15]
    set_col_widths(ws, col_widths)

    row = 1

    # ── 事案メタ情報 ──
    write_section_title(ws, row, f"事案情報: {rd.get('case_title', '')}", NUM_COLS)
    row += 1
    meta_items = [
        ("run_id", rd["run_id"]),
        ("実行日時", rd["run_datetime"]),
        ("モデル", rd.get("model", "gpt-5")),
        ("原告", rd.get("plaintiff", "")),
        ("被告", rd.get("defendant", "")),
        ("最大ラウンド数", str(rd.get("max_rounds", ""))),
        ("実際ラウンド数", str(rd.get("actual_rounds", ""))),
        ("事案概要", rd.get("case_summary", "")),
    ]
    for label, val in meta_items:
        write_cell(ws, row, 1, label, font=LABEL_FONT)
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=NUM_COLS)
        write_cell(ws, row, 2, val)
        _auto_row_height(ws, row, col_widths, ["", str(val)])
        row += 1
    row += 1

    # ── セクション1: 弁論経過 ──
    row = _write_arguments_section(ws, row, rd, col_widths)
    row += 1

    # ── セクション2: 争点整理表 ──
    row = _write_issues_section(ws, row, rd, col_widths)
    row += 1

    # ── セクション3: 合議詳細 ──
    row = _write_deliberation_section(ws, row, rd, col_widths)
    row += 1

    # ── セクション4: 判決詳細 ──
    row = _write_judgment_section(ws, row, rd, col_widths)
    row += 1

    # ── セクション5: 批評・改善 ──
    row = _write_critique_section(ws, row, rd, col_widths)


def _write_arguments_section(ws, row, rd, col_widths):
    """弁論経過セクションを書き込む。"""
    write_section_title(ws, row, "弁論経過", NUM_COLS)
    row += 1

    headers = ["ラウンド", "話者", "争点ID", "争点タイトル", "",
               "主張（claim）", "理由付け（reasoning）", ""]
    write_header(ws, row, headers, col_widths)
    row += 1

    for rs in rd.get("round_summaries", []):
        rn = rs.get("round_number", 0)
        summary = rs.get("summary", "")
        if rn == 0:
            label = f"【初期整理】{summary}"
        else:
            label = f"【ラウンド {rn}】{summary}"
        write_subheader(ws, row, label, NUM_COLS)
        _auto_row_height(ws, row, [sum(col_widths)], [label])
        row += 1

        for turn in rd.get("turns", []):
            if turn.get("round_number") != rn:
                continue
            speaker = "原告代理人" if turn["speaker"] == "plaintiff" else "被告代理人"
            for content in turn.get("contents", []):
                issue_id = content.get("issue_id", "")
                issue_title = _find_issue_title(rd, issue_id)
                values = [
                    rn, speaker, issue_id, issue_title, "",
                    content.get("claim", ""),
                    content.get("reasoning", ""),
                    "",
                ]
                write_row(ws, row, values, col_widths)
                row += 1

    # 裁判官の継続判定
    for jd in rd.get("judge_decisions", []):
        cont = "はい" if jd.get("continue_round") else "いいえ"
        label = f"【裁判官判定（ラウンド{jd.get('round_number', '')}後）】継続: {cont} ― {jd.get('reason', '')}"
        write_subheader(ws, row, label, NUM_COLS)
        _auto_row_height(ws, row, [sum(col_widths)], [label])
        row += 1

    return row


def _write_issues_section(ws, row, rd, col_widths):
    """争点整理表セクションを書き込む。"""
    write_section_title(ws, row, "争点整理表", NUM_COLS)
    row += 1

    headers = ["issue_id", "争点タイトル", "争点の説明", "原告主張（全文）",
               "被告主張（全文）", "争いのない事実", "関連条文", "ステータス"]
    write_header(ws, row, headers, col_widths)
    row += 1

    for iss in rd.get("issues", []):
        values = [
            iss.get("issue_id", ""),
            iss.get("title", ""),
            iss.get("description", ""),
            iss.get("plaintiff_argument", ""),
            iss.get("defendant_argument", ""),
            ", ".join(iss.get("undisputed_facts", [])),
            ", ".join(iss.get("related_statutes", [])),
            iss.get("status", ""),
        ]
        write_row(ws, row, values, col_widths)
        row += 1

    return row


def _write_deliberation_section(ws, row, rd, col_widths):
    """合議詳細セクションを書き込む。"""
    write_section_title(ws, row, "合議詳細", NUM_COLS)
    row += 1

    headers = ["裁判官", "issue_id", "争点タイトル", "判断",
               "", "理由（全文）", "適用条文", ""]
    write_header(ws, row, headers, col_widths)
    row += 1

    for opinion in rd.get("deliberation_opinions", []):
        issue_id = opinion.get("issue_id", "")
        issue_title = _find_issue_title(rd, issue_id)
        values = [
            JUDGE_NAMES.get(opinion.get("speaker", ""), opinion.get("speaker", "")),
            issue_id,
            issue_title,
            opinion.get("decision", ""),
            "",
            opinion.get("reasoning", ""),
            ", ".join(opinion.get("related_statutes", [])),
            "",
        ]
        write_row(ws, row, values, col_widths)
        row += 1

    return row


def _write_judgment_section(ws, row, rd, col_widths):
    """判決詳細セクションを書き込む。"""
    write_section_title(ws, row, "判決詳細", NUM_COLS)
    row += 1

    headers = ["段階", "issue_id", "争点タイトル", "判断",
               "結論（主文）", "理由（全文）", "適用条文", ""]
    write_header(ws, row, headers, col_widths)
    row += 1

    # 草案
    draft = rd.get("draft_judgment", {})
    if draft:
        for di in draft.get("issues", []):
            values = [
                "判決草案",
                di.get("issue_id", ""),
                di.get("issue_title", ""),
                di.get("decision", ""),
                draft.get("conclusion", ""),
                di.get("reasoning", ""),
                ", ".join(di.get("statutes", [])),
                "",
            ]
            write_row(ws, row, values, col_widths)
            row += 1

    # 最終判決
    final = rd.get("final_judgment", {})
    if final:
        for fi in final.get("issues", []):
            values = [
                "最終判決",
                fi.get("issue_id", ""),
                fi.get("issue_title", ""),
                fi.get("decision", ""),
                final.get("conclusion", ""),
                fi.get("reasoning", ""),
                ", ".join(fi.get("statutes", [])),
                "",
            ]
            write_row(ws, row, values, col_widths)
            row += 1

    return row


def _write_critique_section(ws, row, rd, col_widths):
    """批評・改善セクションを書き込む。"""
    write_section_title(ws, row, "批評・改善", NUM_COLS)
    row += 1

    headers = ["No.", "批評者", "対象争点", "批評内容（問題点）",
               "", "批評詳細（理由）", "改善提案", "対応状況"]
    write_header(ws, row, headers, col_widths)
    row += 1

    for idx, crit in enumerate(rd.get("criticisms", []), 1):
        authored = crit.get("authored_by", "")
        values = [
            idx,
            JUDGE_NAMES.get(authored, authored or "判決批評者"),
            crit.get("target_issue_id", ""),
            crit.get("problem", ""),
            "",
            crit.get("reason", ""),
            crit.get("suggestion", ""),
            crit.get("addressed", ""),
        ]
        write_row(ws, row, values, col_widths)
        row += 1

    return row


def _find_issue_title(rd, issue_id):
    for iss in rd.get("issues", []):
        if iss.get("issue_id") == issue_id:
            return iss.get("title", "")
    return ""


# ================================================================
# メイン
# ================================================================
def main():
    runs = collect_runs()
    if not runs:
        print("No run directories found in", OUTPUTS_DIR)
        return

    all_runs_data = [extract_run_data(r) for r in runs]
    # 空のrun（失敗・中断等）を除外
    all_runs_data = [rd for rd in all_runs_data if rd.get("turns") or rd.get("issues")]
    print(f"Found {len(all_runs_data)} run(s) with data")

    wb = Workbook()

    # Sheet 1: 横断比較
    build_overview_sheet(wb, all_runs_data)

    # Sheet 2〜N: 事案ごとに1シート（弁論→争点→合議→判決→批評）
    for case_no, rd in enumerate(all_runs_data, 1):
        build_case_sheet(wb, case_no, rd)
        print(f"  Sheet created: 事案{case_no} ({rd['run_id']})")

    wb.save(str(OUTPUT_FILE))
    print(f"Excel saved: {OUTPUT_FILE}")


if __name__ == "__main__":
    main()
